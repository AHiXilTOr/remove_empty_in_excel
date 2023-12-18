import os, openpyxl, logging, gc, time
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Font
from tqdm import tqdm
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from copy import copy

FILE_EXTENSIONS = ('.xlsx', '.xls')
FILE_PATH = r'C:\Users\s1\Desktop\python-xls'
OUTPUT_FOLDER = r'processed_files'

output_folder = os.path.join(FILE_PATH, OUTPUT_FOLDER)
os.makedirs(output_folder, exist_ok=True)

logging.basicConfig(filename='log.txt', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def remove_empty_rows(sheet):
    '''
    rows = [row[0].row for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row) if not any(cell.value for cell in row)]

    if all(cell.value is None for cell in row)
    
    for row_idx in reversed(rows):
        sheet.delete_rows(row_idx)
    '''
    
    '''
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        if any(cell.value for cell in row):
            continue
    '''

    '''
    for row in rows:
        for cell in row:
            cell.value = None
            cell.border = Border()
            cell.fill = PatternFill()
    '''
    df = pd.DataFrame(sheet.values)
    non_empty_rows = df.dropna(how='all')
    deleted = sheet.max_row
    sheet.delete_rows(1, sheet.max_row)

    for index, row in enumerate(non_empty_rows.itertuples(), start=1):
        sheet.append(list(row[1:]))
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                new_aligment = copy(cell.alignment)
                new_aligment.wrap_text=True
                cell.alignment = new_aligment

            if index == 1:
                for cell in row:
                    cell.font = Font(bold=True)
            else:
                pass
    '''
    rows_to_delete = set()

    for row in range(1, sheet.max_row + 1):
        is_empty = all(sheet.cell(row=row, column=col).value is None for col in range(1, sheet.max_column + 1))
        if is_empty:
            rows_to_delete.add(row)

    for row in reversed(list(rows_to_delete)):
        sheet.delete_rows(row)
    '''
    return deleted

def apply_borders(ws, start_row, end_row):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for row in range(start_row, end_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            
def visualize_total_removal(file_names, initial_row_counts, final_row_counts):
    
    # Выбор минимального, среднего и максимального значения
    '''
    min_index = np.argmin(final_row_counts)
    max_index = np.argmax(final_row_counts)
    mean_index = np.argsort(final_row_counts)[len(final_row_counts)//2]

    file_names = [file_names[min_index], file_names[mean_index], file_names[max_index]]
    initial_counts = [initial_row_counts[min_index], initial_row_counts[mean_index], initial_row_counts[max_index]]
    final_counts = [final_row_counts[min_index], final_row_counts[mean_index],final_row_counts[max_index]]
    '''
    
    x = np.arange(len(file_names))
    bar_width = 0.4

    fig, ax = plt.subplots(figsize=(12,8))

    bars1 = ax.bar(x - bar_width/2, initial_row_counts, bar_width, label='В начале', color='blue', edgecolor='black', linewidth=1, alpha=0.7)
    bars2 = ax.bar(x + bar_width/2, final_row_counts, bar_width, label='В конце', color='green', edgecolor='black', linewidth=1, alpha=0.7)

    ax.set_title('Общее количество рядов')
    ax.set_ylabel('Кол-во рядов')
    ax.set_xticks(x)
    #ax.set_xticklabels(file_names)

    legend = ax.legend()
    frame = legend.get_frame()
    frame.set_facecolor('lightgray')
    frame.set_edgecolor('black')

    for bar1, bar2, initial_count, final_count in zip(bars1, bars2, initial_row_counts, final_row_counts):
        ax.text(bar1.get_x() + bar1.get_width() / 2, bar1.get_height() + 0.02 * max(initial_row_counts), f'{initial_count}', ha='center', va='bottom', color='black', fontsize=8)
        ax.text(bar2.get_x() + bar2.get_width() / 2, bar2.get_height() + 0.02 * max(final_row_counts), f'{final_count}', ha='center', va='bottom', color='black', fontsize=8)

    ax.yaxis.grid(True, linestyle='--', alpha=0.7)

    plt.tight_layout()
    plt.show()

def get_row_counts(file_path, output_folder):
    file_name, file_extension = os.path.splitext(os.path.basename(file_path))
    output_file_path = os.path.join(output_folder, f'{file_name}.xlsx')
    initial_row_count = 0
    final_row_count = 0
    if FLAG:
        df = pd.read_excel(file_path)
        # df = df.dropna(how='all', thresh=1)
        initial_row_count = df.shape[0]
        df = df.dropna(axis=0, how='all')
        '''
        empty_rows = df.apply(lambda row: row.isnull().all(), axis=1)
        df = df.loc[~empty_rows]
        '''
        final_row_count = df.shape[0]
        #df.style.applymap(lambda x: 'border: thin solid border' if pd.notnull(x) else '')
        #df.to_excel(output_file_path, index=False)
        file_path = output_file_path

        # wb = openpyxl.load_workbook(file_path)
        
        '''
        wb = Workbook()
        ws = wb.active
        for r_idx, row in enumerate(df.itertuples(), start=1):
            for c_idx,value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        apply_borders(ws)
        #wb.save(output_file_path)
        '''
        df.to_excel(output_file_path, index=False, header=True)
        gc.collect()
        '''

        logging.info(f"Удалено {initial_row_count-final_row_count} ряда(ов) в {file_name}{file_extension}")
        '''
        #return file_name, initial_row_count, final_row_count
    elif DEFAULT:
        xls = pd.ExcelFile(file_path)

        for sheet_name in xls.sheet_names:
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

            wb = openpyxl.load_workbook(file_path)
            sheet = wb[sheet_name]
            initial_max_row = sheet.max_row
            apply_borders(sheet, start_row=1, end_row=initial_max_row)
            df = df.dropna(axis=0, how='all')

            wb = openpyxl.load_workbook(output_file_path)
            sheet = wb[sheet_name]
            apply_borders(shhet, start_row=1, end_row=initial_max_row)
            df.to_excel(output_file_path, sheet_name=sheet_name, index=False, header=False, engine='openpyxl')
    else:
        if file_extension == '.xls':
            df = pd.read_excel(file_path, header=None)
            df.to_excel(output_file_path, index=False, header=False)
            file_path = output_file_path

        wb = openpyxl.load_workbook(file_path)
            
        for sheet_name in wb.sheetnames:
            initial_row_count += wb[sheet_name].max_row
            deleted = remove_empty_rows(wb[sheet_name])
            final_row_count += initial_row_count - deleted

        logging.info(f"Удалено {initial_row_count-final_row_count} ряда(ов) в {file_name}{file_extension}")
        wb.save(output_file_path)
        wb.close()
        gc.collect()        
        return file_name, initial_row_count, final_row_count
    
def files_in_folder(folder_path, output_folder):
    file_names = []
    initial_row_counts = []
    final_row_counts = []
    
    files = [f for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f)) and f.endswith(FILE_EXTENSIONS)]

    pbar = tqdm(files, unit="file", leave=True, ascii=True)
    for file in pbar:
        pbar.set_description(f"Processing '{file[:10]}...'.{file.split('.')[-1]}")
        file_path = os.path.join(folder_path, file)

        file_name, initial_row_count, final_row_count = get_row_counts(file_path, output_folder)
        #get_row_counts(file_path, output_folder)

        file_names.append(file_name)
        initial_row_counts.append(initial_row_count)
        final_row_counts.append(final_row_count)
        pd.ExcelFile(file_path).close()

    visualize_total_removal(file_names, initial_row_counts, final_row_counts)

FLAG = False
DEFAULT = False
files_in_folder(FILE_PATH, OUTPUT_FOLDER)

