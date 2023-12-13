import os, openpyxl
from openpyxl.styles import PatternFill, Border
from tqdm import tqdm
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np

FILE_EXTENSIONS = (".xlsx", ".xls")
FILE_PATH = r'C:\Users\Админ\Desktop\excel\remove_empty_in_excel'
OUTPUT_FOLDER = r'processed_files'

output_folder = os.path.join(FILE_PATH, OUTPUT_FOLDER)
os.makedirs(output_folder, exist_ok=True)

# Удаления пустых строк
def remove_empty_rows(sheet):
    rows = [row for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row) if not any(cell.value for cell in row)]

    '''Плохо работает'''
    # sheet.delete_rows(row[0].row)

    for row in rows:
        for cell in row:
            cell.border = Border()
            cell.fill = PatternFill()

def visualize_total_removal(file_names, initial_row_counts, final_row_counts):
    x = np.arange(len(file_names))
    bar_width = 0.4

    fig, ax = plt.subplots(figsize=(12, 8))

    bars1 = ax.bar(x - bar_width/2, initial_row_counts, bar_width, label='В начале', color='blue', edgecolor='black', linewidth=1, alpha=0.7)
    bars2 = ax.bar(x + bar_width/2, final_row_counts, bar_width, label='В конце', color='green', edgecolor='black', linewidth=1, alpha=0.7)

    ax.set_title('Общее количество рядов')
    ax.set_ylabel('Кол-во рядов')
    ax.set_xticks(x)
    ax.set_xticklabels(file_names)

    legend = ax.legend()
    frame = legend.get_frame()
    frame.set_facecolor('lightgray') 
    frame.set_edgecolor('black')

    for bar1, bar2, initial_count, final_count in zip(bars1, bars2, initial_row_counts, final_row_counts):
        ax.text(bar1.get_x() + bar1.get_width() / 2, bar1.get_height() + 0.02 * max(initial_row_counts), f'{initial_count}', ha='center', va='bottom', color='black', fontsize=8)
        ax.text(bar2.get_x() + bar2.get_width() / 2, bar2.get_height() + 0.02 * max(final_row_counts), f'{final_count}', ha='center', va='bottom', color='black', fontsize=8)

    ax.yaxis.grid(True, linestyle='--', alpha=0.7)

    plt.tight_layout()
    plt.show()

def files_in_folder(folder_path, output_folder):
    file_names = []
    initial_row_counts = []
    final_row_counts = []

    # Выбор нужного расширения
    files = [f for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f)) and f.endswith(FILE_EXTENSIONS)]

    # Прогресс
    pbar = tqdm(files, unit="file")
    for file in pbar:
        pbar.set_description(f"Processing '{file[:10]}...'.{file.split('.')[-1]}")
        file_path = os.path.join(folder_path, file)

        file_name, initial_row_count, final_row_count = get_row_counts(file_path, output_folder)
        
        file_names.append(file_name)
        initial_row_counts.append(initial_row_count)
        final_row_counts.append(final_row_count)

    # Визуализация общего удаления
    visualize_total_removal(file_names, initial_row_counts, final_row_counts)

def get_row_counts(file_path, output_folder):
    file_name, file_extension = os.path.splitext(os.path.basename(file_path))
    output_file_path = os.path.join(output_folder, f'{file_name}.xlsx')

    if file_extension == '.xls':
        df = pd.read_excel(file_path, header=None)
        df.to_excel(output_file_path, index=False, header=False)
        file_path = output_file_path

    wb = openpyxl.load_workbook(file_path)
    
    initial_row_count = 0
    final_row_count = 0

    for sheet_name in wb.sheetnames:
        initial_row_count += wb[sheet_name].max_row
        remove_empty_rows(wb[sheet_name])
        final_row_count += wb[sheet_name].max_row

    wb.save(output_file_path)
    wb.close()

    return file_name, initial_row_count, final_row_count

files_in_folder(FILE_PATH, output_folder)